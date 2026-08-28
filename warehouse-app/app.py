"""
Memory Block Warehouse — Flask App
"""
import os, io, csv, threading, secrets, logging
from datetime import datetime, timezone, timedelta
from flask import (Flask, jsonify, request, render_template,
                   Response, session, redirect, url_for)
import json as _json
from database import (init_db, get_all_pallets, get_printer_pallet,
                      update_pallet, bulk_update_pallets, get_audit_log,
                      get_stats, get_storage_items, update_storage_item,
                      get_spare_parts, update_spare_part,
                      get_weekly_report_data, search_pallets,
                      log_email_attempt, get_last_email_sent,
                      get_last_email_attempt, get_email_log,
                      save_weekly_snapshot, get_weekly_snapshots,
                      get_printer_items, update_printer_item,
                      get_location_assignments, set_location_assignment, get_l1_pallets,
                      sync_l1_pallets_from_assignments,
                      is_order_processed, log_order, get_order_log, get_pallet_with_above,
                      get_setting, set_setting, log_audit)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
log = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(hours=8)

WAREHOUSE_PIN = os.environ.get('WAREHOUSE_PIN', '2104')

init_db()

# ── Auth helpers ───────────────────────────────────────────────────────────────
def is_authenticated():
    return session.get('auth') == True

def require_auth(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not is_authenticated():
            return jsonify({'error': 'Unauthorised'}), 401
        return f(*args, **kwargs)
    return decorated

# ── Weekly email scheduler ─────────────────────────────────────────────────────
# Polls every 5 minutes instead of sleeping until a computed deadline.
# Persists last-sent to the DB so Railway container restarts cannot skip a week.

def _try_weekly_send():
    now = datetime.now(timezone.utc)

    # Already sent successfully this week?
    last_success = get_last_email_sent()
    if last_success:
        ls_dt = datetime.strptime(last_success, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
        if (now - ls_dt).total_seconds() < 6 * 24 * 3600:
            log.info(f'Weekly email already sent at {last_success} UTC — skipping')
            return

    # Attempted (success or fail) within the last 25 min? Avoid rapid retries.
    last_attempt = get_last_email_attempt()
    if last_attempt:
        la_dt = datetime.strptime(last_attempt, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
        if (now - la_dt).total_seconds() < 1500:
            log.info(f'Last attempt was {last_attempt} UTC — waiting before retry')
            return

    log.info(f'Firing weekly report at {now.strftime("%Y-%m-%d %H:%M:%S")} UTC')
    from email_report import send_weekly_report
    success, code, detail = send_weekly_report()
    log_email_attempt('success' if success else 'failure', code, detail)
    if success:
        log.info(f'Weekly report sent OK (HTTP {code}): {detail}')
        try:
            save_weekly_snapshot()
            log.info('Weekly inventory snapshot saved')
        except Exception as snap_err:
            log.error(f'Snapshot save failed: {snap_err}')
    else:
        log.error(f'Weekly report FAILED (HTTP {code}): {detail}')


def _scheduler():
    import time as _t
    log.info('Scheduler started — polling every 5 min (email window: Monday 23:xx UTC + Unleashed orders)')
    while True:
        _t.sleep(300)
        # ── Unleashed order poll (every tick) ──────────────────────────────
        try:
            _poll_unleashed()
        except Exception as e:
            log.error(f'Unleashed poll error: {e}')
        # ── Weekly email (Monday 23:xx UTC = Tuesday 09:00 AEST) ──────────
        try:
            now = datetime.now(timezone.utc)
            if now.weekday() == 0 and now.hour == 23:
                _try_weekly_send()
        except Exception as e:
            log.error(f'Email scheduler error: {e}')


threading.Thread(target=_scheduler, daemon=True).start()

# ── Pages ──────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if not is_authenticated():
        return render_template('pin.html')
    return render_template('index.html')

# ── PIN auth ───────────────────────────────────────────────────────────────────
@app.route('/api/auth/pin', methods=['POST'])
def api_pin():
    data = request.get_json() or {}
    pin  = str(data.get('pin', ''))
    if pin == WAREHOUSE_PIN:
        session.permanent = True
        session['auth'] = True
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Incorrect PIN'}), 401

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})

# ── Pallets ────────────────────────────────────────────────────────────────────
@app.route('/api/pallets')
@require_auth
def api_pallets():
    return jsonify(get_all_pallets())

@app.route('/api/pallets/<pid>', methods=['PATCH'])
@require_auth
def api_update_pallet(pid):
    data = request.get_json() or {}
    allowed = {'fill','units','name','sku','sku2','sku3','notes','max_units','units_per_box','refurbished_units','pallet_label','units2','units_per_box2','units3','units_per_box3'}
    updates = {k:v for k,v in data.items() if k in allowed}
    if 'fill'          in updates: updates['fill']          = max(0, min(100, int(updates['fill'])))
    if 'units'         in updates: updates['units']         = max(0, int(updates['units']))
    if 'max_units'     in updates: updates['max_units']     = max(0, int(updates['max_units']))
    if 'units_per_box'     in updates: updates['units_per_box']     = max(0, int(updates['units_per_box']))
    if 'refurbished_units' in updates: updates['refurbished_units'] = max(0, int(updates['refurbished_units']))
    # When fill reaches 100% with a known box count, lock that in as capacity
    if updates.get('fill') == 100 and updates.get('units', 0) > 0:
        updates['max_units'] = updates['units']
    updated = update_pallet(pid, updates, data.get('changed_by','Stock User'))
    return jsonify({'success':True,'pallet':updated}) if updated else (jsonify({'error':'Not found'}),404)

@app.route('/api/pallets/bulk', methods=['PATCH'])
@require_auth
def api_bulk():
    data = request.get_json() or {}
    allowed = {'fill','units','name','sku','notes'}
    updates = {k:v for k,v in data.items() if k in allowed}
    if 'fill'  in updates: updates['fill']  = max(0,min(100,int(updates['fill'])))
    if 'units' in updates: updates['units'] = int(updates['units'])
    count = bulk_update_pallets(data.get('ids',[]), updates, data.get('changed_by','Bulk'))
    return jsonify({'success':True,'updated':count})

# ── Storage room ───────────────────────────────────────────────────────────────
@app.route('/api/storage')
@require_auth
def api_storage():
    return jsonify(get_storage_items())

@app.route('/api/storage/<int:slot>', methods=['PATCH'])
@require_auth
def api_update_storage(slot):
    data = request.get_json() or {}
    updated = update_storage_item(slot, data)
    return jsonify({'success':True,'item':updated})

# ── Printer items ──────────────────────────────────────────────────────────────
@app.route('/api/printer-items')
@require_auth
def api_printer_items():
    return jsonify(get_printer_items())

@app.route('/api/printer-items/<int:row_num>', methods=['PATCH'])
@require_auth
def api_update_printer_item(row_num):
    data = request.get_json() or {}
    updated = update_printer_item(row_num, data)
    return jsonify({'success': True, 'item': updated}) if updated else (jsonify({'error': 'Not found'}), 404)

# ── Spare parts ────────────────────────────────────────────────────────────────
@app.route('/api/spare-parts')
@require_auth
def api_spare_parts():
    return jsonify(get_spare_parts())

@app.route('/api/spare-parts/<int:slot>', methods=['PATCH'])
@require_auth
def api_update_spare_part(slot):
    data = request.get_json() or {}
    updated = update_spare_part(slot, data)
    return jsonify({'success':True,'item':updated})

# ── Stats & audit ──────────────────────────────────────────────────────────────
@app.route('/api/stats')
@require_auth
def api_stats():
    return jsonify(get_stats())

@app.route('/api/audit')
@require_auth
def api_audit():
    return jsonify(get_audit_log(int(request.args.get('limit',150))))

# ── Export ─────────────────────────────────────────────────────────────────────
@app.route('/api/export/xlsx')
@require_auth
def api_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from database import SECTION_SEGMENTS
    pallets        = get_all_pallets()
    storage        = get_storage_items()
    spare_parts    = get_spare_parts()
    printer_items  = get_printer_items()

    HDR_FONT  = Font(bold=True, color='F5C800', size=11)
    HDR_FILL  = PatternFill('solid', fgColor='1A1200')
    HDR_ALIGN = Alignment(horizontal='center', vertical='center')
    FILL_FULL  = PatternFill('solid', fgColor='C8F0D5')
    FILL_MED   = PatternFill('solid', fgColor='FFF9C4')
    FILL_LOW   = PatternFill('solid', fgColor='FFD5D3')
    FILL_EMPTY = PatternFill('solid', fgColor='EEEEEE')
    thin = Side(style='thin', color='DDDDDD')
    BORDER = Border(bottom=Side(style='thin', color='EEEEEE'))

    def apply_header(ws, headers, widths):
        ws.row_dimensions[1].height = 22
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(1, c, h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'A2'

    def status_fill(f):
        if f == 0:  return FILL_EMPTY
        if f < 30:  return FILL_LOW
        if f < 60:  return FILL_MED
        return FILL_FULL

    wb = Workbook()

    # ── Sheet 1: Rack Pallets ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Rack Pallets'
    hdrs = ['Pallet ID','Section','Segment','Slot','Level','Product','SKU',
            'Fill %','Boxes','Units/Box','Total Units','Refurb Units','Status','Notes']
    wids = [16, 10, 10, 8, 8, 28, 16, 10, 10, 12, 14, 14, 12, 30]
    apply_header(ws1, hdrs, wids)
    rack_pallets = [p for p in pallets if not p.get('is_printer')]
    for row_i, p in enumerate(rack_pallets, 2):
        f   = p['fill']
        sec = p['section']
        seg_num = SECTION_SEGMENTS.get(sec, p['segment']+1) - p['segment']
        boxes   = p['units']
        upb     = int(p.get('units_per_box') or 0)
        total_u = boxes * upb if upb else ''
        refurb  = int(p.get('refurbished_units') or 0)
        status  = 'Empty' if f==0 else 'Low' if f<30 else 'Medium' if f<60 else 'Full/High'
        vals = [p['id'], sec, seg_num, p['slot']+1, p['level']+1,
                p['name'], p['sku'], f, boxes, upb or '', total_u, refurb or '', status,
                p.get('notes','')]
        sf = status_fill(f)
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row_i, c, v)
            cell.border = BORDER
            if c in (8, 13):  # fill % and status columns
                cell.fill = sf
                cell.alignment = Alignment(horizontal='center')

    # ── Sheet 2: Printer Tracker ───────────────────────────────────────────────
    ws2 = wb.create_sheet('Printer Tracker')
    apply_header(ws2, ['Row', 'Printer ID', 'Laptop ID', 'Notes'], [8, 24, 24, 50])
    for r, item in enumerate(printer_items, 2):
        for c, v in enumerate([item['row_num'], item['printer_id'], item['laptop_id'], item['notes']], 1):
            ws2.cell(r, c, v).border = BORDER

    # ── Sheet 3: Stock Summary by SKU ─────────────────────────────────────────
    ws3 = wb.create_sheet('Stock Summary')
    apply_header(ws3, ['SKU','Product','Pallets','Total Boxes','Units/Box','Total Units','Refurb Units','Avg Fill %'], [18,30,10,14,12,14,14,12])
    # Group non-printer pallets by SKU
    from collections import defaultdict
    sku_groups = defaultdict(lambda: {'name':'','pallets':0,'boxes':0,'upb_set':set(),'fill_sum':0,'refurb':0})
    for p in rack_pallets:
        sku = p['sku'] or '(no SKU)'
        g   = sku_groups[sku]
        g['name']     = g['name'] or p['name']
        g['pallets'] += 1
        g['boxes']   += int(p['units'] or 0)
        g['fill_sum']+= int(p['fill'] or 0)
        g['refurb']  += int(p.get('refurbished_units') or 0)
        upb = int(p.get('units_per_box') or 0)
        if upb: g['upb_set'].add(upb)
    for r_i, (sku, g) in enumerate(sorted(sku_groups.items()), 2):
        upb_vals = sorted(g['upb_set'])
        upb_str  = '/'.join(str(v) for v in upb_vals) if upb_vals else ''
        upb_num  = upb_vals[0] if len(upb_vals)==1 else ''
        total_u  = g['boxes'] * upb_num if upb_num else ''
        avg_fill = round(g['fill_sum'] / g['pallets']) if g['pallets'] else 0
        refurb   = g['refurb'] if g['refurb'] else ''
        vals = [sku, g['name'], g['pallets'], g['boxes'], upb_str, total_u, refurb, avg_fill]
        sf = status_fill(avg_fill)
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(r_i, c, v)
            cell.border = BORDER
            if c == 8:
                cell.fill = sf
                cell.alignment = Alignment(horizontal='center')

    # ── Sheet 5: Storage Room ─────────────────────────────────────────────────
    ws5 = wb.create_sheet('Storage Room')
    apply_header(ws5, ['Slot','Product Name','SKU','Quantity','Notes'], [8,30,18,12,40])
    for r, s in enumerate(storage, 2):
        for c, v in enumerate([s['slot'], s['name'], s['sku'], s['quantity'], s.get('notes','')], 1):
            ws5.cell(r, c, v).border = BORDER

    # ── Sheet 6: Spare Parts ──────────────────────────────────────────────────
    ws6 = wb.create_sheet('Spare Parts')
    apply_header(ws6, ['Slot','Part Name','SKU','Quantity','Notes'], [8,30,18,12,40])
    for r, sp in enumerate(spare_parts, 2):
        for c, v in enumerate([sp['slot'], sp['name'], sp['sku'], sp['quantity'], sp.get('notes','')], 1):
            ws6.cell(r, c, v).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now().strftime('%Y-%m-%d')
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=mb-warehouse-{now}.xlsx'})

# ── Search ─────────────────────────────────────────────────────────────────────
@app.route('/api/search')
@require_auth
def api_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    return jsonify(search_pallets(q))

# ── Dashboard data ─────────────────────────────────────────────────────────────
@app.route('/api/dashboard')
@require_auth
def api_dashboard():
    data = get_weekly_report_data()
    data['stats']         = get_stats()
    data['printer_items'] = get_printer_items()
    return jsonify(data)

@app.route('/dashboard')
def dashboard():
    if not is_authenticated():
        return redirect(url_for('index'))
    return render_template('dashboard.html')

# ── Manual / test email send ───────────────────────────────────────────────────
@app.route('/api/send-report', methods=['POST'])
def api_send_report():
    # Accept either browser session auth OR a CRON_SECRET header (for external cron)
    cron_secret  = os.environ.get('CRON_SECRET', '')
    auth_header  = request.headers.get('X-Cron-Secret', '')
    if not is_authenticated() and not (cron_secret and auth_header == cron_secret):
        return jsonify({'error': 'Unauthorised'}), 401

    try:
        from email_report import send_weekly_report
        success, code, detail = send_weekly_report()
        log_email_attempt('success' if success else 'failure', code, detail)
        log.info(f'Manual send: HTTP {code} — {detail}')
        return jsonify({'success': success, 'http_code': code, 'detail': detail})
    except Exception as e:
        log.error(f'Manual send error: {e}')
        return jsonify({'success': False, 'http_code': 0, 'detail': str(e)}), 500


@app.route('/api/email-log')
@require_auth
def api_email_log():
    return jsonify(get_email_log())


@app.route('/api/snapshots')
@require_auth
def api_snapshots():
    return jsonify(get_weekly_snapshots())

# ── Location assignments ────────────────────────────────────────────────────────
@app.route('/locations')
def locations_page():
    if not is_authenticated():
        return redirect(url_for('index'))
    return render_template('locations.html')

@app.route('/sheet')
def sheet_page():
    if not is_authenticated():
        return redirect(url_for('index'))
    return render_template('sheet.html')

@app.route('/api/locations')
@require_auth
def api_locations():
    return jsonify({
        'pallets':     get_l1_pallets(),
        'assignments': get_location_assignments(),
    })

@app.route('/api/locations/<pid>', methods=['PATCH'])
@require_auth
def api_set_location(pid):
    data          = request.get_json() or {}
    sku           = data.get('sku', '')
    product_name  = data.get('product_name', '')
    sku2          = data.get('sku2', '')
    product_name2 = data.get('product_name2', '')
    sku3          = data.get('sku3', '')
    product_name3 = data.get('product_name3', '')
    assigned_by   = data.get('assigned_by', 'Staff')
    set_location_assignment(pid, sku, product_name, assigned_by,
                            sku2, product_name2, sku3, product_name3)
    return jsonify({'success': True})

# ── Shared order-processing logic ─────────────────────────────────────────────

def _process_order(order_num: str, customer: str, lines: list, raw_str: str,
                   source: str = 'Unleashed'):
    """
    Deduct stock for one completed sales order.
    Returns (processed, skipped, alerts_fired).
    Logs the order to order_log when done.
    """
    # Build SKU lookup: rack L1 pallets
    assignments = get_location_assignments()
    sku_to_pallet = {}
    for pid, a in assignments.items():
        for sku_field in ('sku', 'sku2', 'sku3'):
            s = a.get(sku_field, '').strip()
            if s:
                sku_to_pallet[s.upper()] = pid

    # Build SKU lookup: spare parts  {sku_upper: slot}
    sku_to_spare = {}
    for item in get_spare_parts():
        s = (item.get('sku') or '').strip()
        if s:
            sku_to_spare[s.upper()] = item['slot']

    # Build SKU lookup: storage room  {sku_upper: slot}
    sku_to_storage = {}
    for item in get_storage_items():
        s = (item.get('sku') or '').strip()
        if s:
            sku_to_storage[s.upper()] = item['slot']

    processed = 0
    skipped   = 0
    skip_notes= []
    restock_alerts = []

    for line in lines:
        product   = line.get('Product') or {}
        sku       = (product.get('ProductCode') or product.get('productCode') or '').strip()
        qty_units = float(line.get('OrderQuantity') or line.get('orderQuantity') or 0)

        if not sku or qty_units <= 0:
            skipped += 1; continue

        sku_upper = sku.upper()
        prod_name = product.get('ProductDescription', sku)

        # ── Try rack L1 pallet first ──────────────────────────────────────────
        pid = sku_to_pallet.get(sku_upper)
        if not pid and sku_upper.endswith('BOX'):
            pid = sku_to_pallet.get(sku_upper[:-3])

        if pid:
            pallet, above = get_pallet_with_above(pid)
            if not pallet:
                skip_notes.append(f'{sku}: pallet {pid} not found')
                skipped += 1; continue

            upb = int(pallet.get('units_per_box') or 0)
            if upb <= 0:
                skip_notes.append(f'{sku} ({pid}): units_per_box not set')
                skipped += 1; continue

            boxes_to_deduct = round(qty_units / upb)
            if boxes_to_deduct <= 0:
                skipped += 1; continue

            current_units = int(pallet.get('units') or 0)
            max_units     = int(pallet.get('max_units') or 0)
            new_units     = max(0, current_units - boxes_to_deduct)

            updates = {'units': new_units}
            if max_units > 0:
                updates['fill'] = round(new_units / max_units * 100)
            elif current_units > 0:
                updates['fill'] = round(new_units / current_units * (pallet.get('fill') or 0))
            updates['fill'] = max(0, min(100, updates.get('fill', pallet.get('fill', 0))))

            update_pallet(pid, updates, f'{source} Order {order_num}')
            processed += 1

            if updates['fill'] < 30:
                restock_alerts.append({
                    'pallet_id':     pid,
                    'sku':           sku,
                    'product_name':  prod_name,
                    'new_units':     new_units,
                    'new_fill':      updates['fill'],
                    'order_number':  order_num,
                    'customer':      customer,
                    'above_pallets': above,
                })
            continue

        # ── Try spare parts ───────────────────────────────────────────────────
        spare_slot = sku_to_spare.get(sku_upper)
        if spare_slot:
            current_item = next((i for i in get_spare_parts() if i['slot'] == spare_slot), {})
            old_qty  = int(current_item.get('quantity', 0))
            new_qty  = max(0, old_qty - int(qty_units))
            updated  = update_spare_part(spare_slot, {'quantity': new_qty})
            if updated:
                log_audit(f'SPARE-{spare_slot} ({sku})', f'{source} Order {order_num}',
                          'quantity', old_qty, new_qty)
                processed += 1
                if new_qty == 0:
                    restock_alerts.append({
                        'pallet_id':     f'Spare Parts slot {spare_slot}',
                        'sku':           sku,
                        'product_name':  prod_name,
                        'new_units':     0,
                        'new_fill':      0,
                        'order_number':  order_num,
                        'customer':      customer,
                        'above_pallets': [],
                    })
            continue

        # ── Try storage room ──────────────────────────────────────────────────
        storage_slot = sku_to_storage.get(sku_upper)
        if storage_slot:
            current_item = next((i for i in get_storage_items() if i['slot'] == storage_slot), {})
            old_qty  = int(current_item.get('quantity', 0))
            new_qty  = max(0, old_qty - int(qty_units))
            updated  = update_storage_item(storage_slot, {'quantity': new_qty})
            if updated:
                log_audit(f'STORAGE-{storage_slot} ({sku})', f'{source} Order {order_num}',
                          'quantity', old_qty, new_qty)
                processed += 1
                if new_qty == 0:
                    restock_alerts.append({
                        'pallet_id':     f'Storage Room slot {storage_slot}',
                        'sku':           sku,
                        'product_name':  prod_name,
                        'new_units':     0,
                        'new_fill':      0,
                        'order_number':  order_num,
                        'customer':      customer,
                        'above_pallets': [],
                    })
            continue

        skip_notes.append(f'{sku}: no location assigned')
        skipped += 1

    notes_str    = '; '.join(skip_notes) if skip_notes else ''
    alerts_fired = 0
    if restock_alerts:
        from email_report import send_restock_alert
        ok, code, detail = send_restock_alert(restock_alerts)
        alerts_fired = len(restock_alerts) if ok else 0
        log.info(f'Restock alert ({order_num}): {detail} (HTTP {code})')

    log_order(order_num, customer, processed, skipped, alerts_fired, notes_str, raw_str[:4000])
    log.info(f'{source} order {order_num}: {processed} processed, {skipped} skipped, {alerts_fired} alerts')
    return processed, skipped, alerts_fired


# ── Unleashed webhook (kept as fallback) ──────────────────────────────────────
@app.route('/api/webhooks/unleashed', methods=['POST'])
def api_unleashed_webhook():
    secret   = os.environ.get('WEBHOOK_SECRET', '')
    provided = request.args.get('secret') or request.headers.get('X-Webhook-Secret', '')
    if secret and provided != secret:
        log.warning('Unleashed webhook: invalid secret')
        return jsonify({'error': 'Forbidden'}), 403

    raw = request.get_data(as_text=True)
    try:
        data = _json.loads(raw)
    except Exception:
        return jsonify({'error': 'Invalid JSON'}), 400

    event = data.get('Event', '')
    if event not in ('SalesOrderComplete', 'SalesOrderCompleted'):
        return jsonify({'ignored': True, 'event': event}), 200

    order     = data.get('SalesOrder') or data.get('salesOrder') or {}
    order_num = order.get('OrderNumber') or order.get('orderNumber', 'UNKNOWN')
    customer  = (order.get('Customer') or {}).get('CustomerName', '')
    lines     = order.get('SalesOrderLines') or order.get('salesOrderLines') or []

    if is_order_processed(order_num):
        log.info(f'Webhook: order {order_num} already processed — skipping')
        return jsonify({'skipped': True, 'reason': 'already processed'}), 200

    processed, skipped, alerts = _process_order(order_num, customer, lines, raw, 'Webhook')
    return jsonify({'success': True, 'order': order_num,
                    'processed': processed, 'skipped': skipped, 'alerts': alerts}), 200


# ── Unleashed API poller ───────────────────────────────────────────────────────

def _poll_unleashed():
    """Fetch completed orders from Unleashed API since last poll and process new ones."""
    api_id  = os.environ.get('UNLEASHED_API_ID',  '').strip()
    api_key = os.environ.get('UNLEASHED_API_KEY', '').strip()
    if not api_id or not api_key:
        return  # credentials not configured — skip silently

    from unleashed import get_completed_orders

    # Use last-poll date stored in DB; default to 7 days ago on first run
    from datetime import timedelta as _td
    last_poll = get_setting('unleashed_last_poll')
    if last_poll:
        # Re-query from 2 days before last poll to catch any late-completing orders
        try:
            last_dt = datetime.strptime(last_poll[:10], '%Y-%m-%d').date()
            start   = (last_dt - _td(days=2)).strftime('%Y-%m-%d')
        except Exception:
            start = (datetime.now(timezone.utc).date() - _td(days=7)).strftime('%Y-%m-%d')
    else:
        start = (datetime.now(timezone.utc).date() - _td(days=7)).strftime('%Y-%m-%d')

    now_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

    try:
        items = get_completed_orders(start)
        log.info(f'Unleashed poll: {len(items)} completed orders since {start}')
    except Exception as e:
        log.error(f'Unleashed poll failed: {e}')
        return

    new_orders  = 0
    skipped_wh  = 0
    import json as _j
    for order in items:
        order_num = order.get('OrderNumber', '')
        if not order_num or is_order_processed(order_num):
            continue

        # Only process orders dispatched from the Cheltenham warehouse (WHAU)
        wh_code = (order.get('Warehouse') or {}).get('WarehouseCode', '').strip()
        if wh_code and wh_code != 'WHAU':
            log.info(f'Skipping order {order_num} — warehouse {wh_code} (not WHAU)')
            skipped_wh += 1
            continue

        customer = (order.get('Customer') or {}).get('CustomerName', '')
        lines    = order.get('SalesOrderLines') or []
        _process_order(order_num, customer, lines, _j.dumps(order), 'Unleashed Poll')
        new_orders += 1

    if skipped_wh:
        log.info(f'Unleashed poll: skipped {skipped_wh} non-AU order(s)')

    set_setting('unleashed_last_poll', now_str)
    if new_orders:
        log.info(f'Unleashed poll: processed {new_orders} new order(s)')


# ── Manual poll trigger (for testing) ─────────────────────────────────────────
@app.route('/api/unleashed/poll', methods=['POST'])
@require_auth
def api_trigger_poll():
    try:
        # Always look back 7 days on manual trigger but never reprocess already-logged orders
        set_setting('unleashed_last_poll', '')
        _poll_unleashed()
        return jsonify({'success': True, 'last_poll': get_setting('unleashed_last_poll')})
    except Exception as e:
        log.error(f'Manual poll error: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/unleashed/status')
@require_auth
def api_unleashed_status():
    configured = bool(os.environ.get('UNLEASHED_API_ID') and os.environ.get('UNLEASHED_API_KEY'))
    return jsonify({
        'configured':  configured,
        'last_poll':   get_setting('unleashed_last_poll'),
        'order_count': len(get_order_log(1000)),
    })


@app.route('/api/locations/sync', methods=['POST'])
@require_auth
def api_sync_locations():
    updated = sync_l1_pallets_from_assignments()
    log.info(f'Location sync: updated {updated} L1 pallets')
    return jsonify({'success': True, 'updated': updated})

@app.route('/api/order-log')
@require_auth
def api_order_log():
    return jsonify(get_order_log())


@app.route('/health')
def health():
    return jsonify({'status': 'ok'}), 200


# ── Google Sheets sync ────────────────────────────────────────────────────────

@app.route('/api/sheets/push', methods=['POST'])
@require_auth
def api_sheets_push():
    try:
        from google_sheets import push_to_sheet
        pallets = get_all_pallets()
        count   = push_to_sheet(pallets)
        return jsonify({'success': True, 'rows': count})
    except Exception as e:
        log.error(f'Sheets push failed: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sheets/pull', methods=['POST'])
@require_auth
def api_sheets_pull():
    try:
        from google_sheets import pull_from_sheet
        updates = pull_from_sheet()
        allowed = {'pallet_label','name','sku','sku2','sku3','fill','units',
                   'units_per_box','units2','units_per_box2','units3',
                   'units_per_box3','notes','refurbished_units'}
        ok = 0
        for row in updates:
            pid     = row.pop('pid', '')
            cleaned = {k: v for k, v in row.items() if k in allowed}
            if pid and update_pallet(pid, cleaned, 'Google Sheet Sync'):
                ok += 1
        return jsonify({'success': True, 'updated': ok, 'total': len(updates)})
    except Exception as e:
        log.error(f'Sheets pull failed: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT',5000))
    print(f'\n  Memory Block Warehouse: http://localhost:{port}\n')
    app.run(host='0.0.0.0', port=port, debug=False)
